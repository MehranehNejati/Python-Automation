from openpyxl import load_workbook, Workbook

def generate_sales_report(input_file="sales.xlsx", output_file="sales_report.xlsx"):
    # Load workbook
    wb = load_workbook(input_file)
    sheet = wb.active
    
    # Extract data
    totals = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        item, price, quantity = row[0], row[1], row[2]
        total = price * quantity
        totals[item] = totals.get(item, 0) + total
        
    # Create new workbook for report
    report_wb = Workbook()
    report_sheet = report_wb.active
    report_sheet.title = "Sales Summary"
    
    # Write headers
    report_sheet.append(["Item", "Total Sales"])
    
    # Write totals
    for item, total in totals.items():
        report_sheet.append([item, total])
        
    # Save report
    report_wb.save(output_file)
    print(f"Report saved to {output_file}")

    # Run it
if __name__ == "__main__":
    generate_sales_report()